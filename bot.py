import os
import logging
import requests
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes, ConversationHandler
import json
from datetime import datetime, timedelta
import math
from openpyxl import load_workbook
import asyncio
from flask import Flask
from threading import Thread
import time

# Replit keep-alive server
app = Flask('')

@app.route('/')
def home():
    return "🚀 Telegram Stair Bot is Alive and Running!"

@app.route('/ping')
def ping():
    return "PONG"

@app.route('/status')
def status():
    return {
        "status": "active",
        "timestamp": datetime.now().isoformat(),
        "service": "telegram-stair-bot"
    }

def run_flask():
    app.run(host='0.0.0.0', port=8080)

def keep_alive():
    """Запускает Flask сервер в отдельном потоке"""
    t = Thread(target=run_flask)
    t.daemon = True
    t.start()
    logging.info("🔄 Keep-alive server started on port 8080")

def start_ping_loop():
    """Фоновая задача для само-пинга (опционально)"""
    def ping_loop():
        while True:
            try:
                # Получаем URL Replit из переменных окружения
                repl_url = os.getenv('REPLIT_URL')
                if repl_url:
                    requests.get(f"{repl_url}/ping", timeout=10)
                    logging.debug("🔁 Self-ping completed")
            except Exception as e:
                logging.debug(f"🔁 Self-ping failed: {e}")
            time.sleep(300)  # Пинг каждые 5 минут
    
    t = Thread(target=ping_loop)
    t.daemon = True
    t.start()

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Состояния диалога
SELECTING_TYPE, SELECTING_CONFIG, INPUT_HEIGHT, SELECTING_STEP_SIZE, SEARCH_MATERIAL = range(5)

# Глобальные переменные для хранения данных
user_data = {}
prices_data = None
last_price_update = None
PRICE_UPDATE_INTERVAL = timedelta(hours=24)
MESSAGES_TO_DELETE = {}

# Константы расчета
FIXED_STEP_HEIGHT = 225
MAX_STRINGER_LENGTH = 4000

def load_prices(force_update=False):
    """Загрузка цен из Excel файла с автообновлением"""
    global prices_data, last_price_update
    
    try:
        current_time = datetime.now()
        if force_update or last_price_update is None or (current_time - last_price_update) > PRICE_UPDATE_INTERVAL:
            logger.info("Начинаем обновление цен...")
            
            wb = load_workbook('data.xlsx', data_only=True)
            sheet = wb.active
            
            prices = []
            
            for row_num in range(4, sheet.max_row + 1):
                article = sheet.cell(row=row_num, column=1).value
                name = sheet.cell(row=row_num, column=2).value
                stair_type = sheet.cell(row=row_num, column=3).value
                sizes = sheet.cell(row=row_num, column=4).value
                unit = sheet.cell(row=row_num, column=5).value
                price = sheet.cell(row=row_num, column=6).value
                
                if article and name and price:
                    item = {
                        'article': str(article).split('.')[0] if '.' in str(article) else str(article),
                        'name': str(name),
                        'stair_type': str(stair_type) if stair_type else '',
                        'sizes': str(sizes) if sizes else '',
                        'unit': str(unit) if unit else 'шт.',
                        'price': float(price) if price else 0
                    }
                    prices.append(item)
            
            prices_data = prices
            last_price_update = current_time
            logger.info(f"Успешно загружено {len(prices)} позиций из Excel")
        else:
            logger.info("Используем кэшированные цены")
            
    except Exception as e:
        logger.error(f"Ошибка загрузки прайса: {e}")
        prices_data = get_test_data()

def get_test_data():
    """Тестовые данные если файл не загружается"""
    test_data = [
        {'article': '15762294', 'name': 'Верхний и нижний элемент сталь ЛЭ-01-01', 'stair_type': 'металлическая', 'price': 7590, 'unit': 'штука'},
        {'article': '15762307', 'name': 'Промежуточный элемент сталь ЛЭ-01-02', 'stair_type': 'металлическая', 'price': 4076, 'unit': 'штука'},
        {'article': '15762374', 'name': 'Опора лестницы 1000мм сталь ЛЭ-01-09', 'stair_type': 'металлическая', 'price': 3647, 'unit': 'штука'},
        {'article': '15762382', 'name': 'Опора лестницы 2000 сталь ЛЭ-01-10', 'stair_type': 'металлическая', 'price': 5490, 'unit': 'штука'},
        {'article': '15762391', 'name': 'Угловой элемент сталь ЛЭ-01-14', 'stair_type': 'металлическая', 'price': 12411, 'unit': 'штука'},
        {'article': '83850952', 'name': 'СТУПЕНЬ ПРЯМАЯ 900x300', 'stair_type': 'деревянная', 'price': 1504, 'unit': 'штука'},
        {'article': '83850953', 'name': 'СТУПЕНЬ ПРЯМАЯ 1000x300', 'stair_type': 'деревянная', 'price': 1282, 'unit': 'штука'},
        {'article': '83850954', 'name': 'СТУПЕНЬ ПРЯМАЯ 1200x300', 'stair_type': 'деревянная', 'price': 1358, 'unit': 'штука'},
        {'article': '83850961', 'name': 'Тетива 3000x300x60', 'stair_type': 'деревянная', 'price': 9518, 'unit': 'штука'},
        {'article': '83850962', 'name': 'Тетива 4000x300x60', 'stair_type': 'деревянная', 'price': 10215, 'unit': 'штука'},
        {'article': '83850939', 'name': 'Поручень 3000мм', 'stair_type': 'деревянная', 'price': 2108, 'unit': 'штука'},
        {'article': '89426866', 'name': 'Столб Хюгге', 'stair_type': 'деревянная', 'price': 1931, 'unit': 'штука'},
        {'article': '89426868', 'name': 'Балясина Хюгге', 'stair_type': 'деревянная', 'price': 400, 'unit': 'штука'},
        {'article': 'platform_1000', 'name': 'Площадка 1000x1000', 'stair_type': 'деревянная', 'price': 8000, 'unit': 'штука'},
        {'article': 'platform_1200', 'name': 'Площадка 1200x1200', 'stair_type': 'деревянная', 'price': 9500, 'unit': 'штука'},
    ]
    logger.info("Используются тестовые данные")
    return test_data

def get_material_price(material_type, name_pattern, default_price):
    """Получение цены с фильтрацией по типу лестницы"""
    if not prices_data:
        return default_price
    
    try:
        for item in prices_data:
            if (item['stair_type'] == material_type and 
                name_pattern.lower() in item['name'].lower()):
                return item['price']
        return default_price
    except Exception as e:
        logger.error(f"Ошибка поиска цены: {e}")
        return default_price

def get_material_by_article(article):
    """Получение материала по артикулу"""
    if not prices_data:
        return None
    
    try:
        clean_article = str(article).split('.')[0] if '.' in str(article) else str(article)
        for item in prices_data:
            if item['article'] == clean_article:
                return item
        return None
    except Exception as e:
        logger.error(f"Ошибка поиска по артикулу {article}: {e}")
        return None

def search_materials_by_article_or_name(search_term):
    """Поиск материалов по артикулу или названию"""
    if not prices_data:
        return []
    
    try:
        search_term = search_term.lower().strip()
        results = []
        
        for item in prices_data:
            if (search_term in item['article'].lower() or 
                search_term in item['name'].lower()):
                results.append(item)
        
        return results
    except Exception as e:
        logger.error(f"Ошибка поиска материалов: {e}")
        return []

def validate_input(value, min_val, max_val, field_name):
    """Проверка ввода на адекватность"""
    try:
        num = float(value)
        if min_val <= num <= max_val:
            return True, num
        else:
            return False, f"❌ {field_name} должен быть от {min_val} до {max_val} мм"
    except ValueError:
        return False, "❌ Пожалуйста, введите число"

async def add_message_to_delete(chat_id, message_id):
    """Добавляем сообщение в список для удаления"""
    if chat_id not in MESSAGES_TO_DELETE:
        MESSAGES_TO_DELETE[chat_id] = []
    MESSAGES_TO_DELETE[chat_id].append(message_id)
    
    if len(MESSAGES_TO_DELETE[chat_id]) > 50:
        MESSAGES_TO_DELETE[chat_id] = MESSAGES_TO_DELETE[chat_id][-50:]

async def cleanup_chat_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Очистка истории чата"""
    try:
        chat_id = update.effective_chat.id
        
        if chat_id in MESSAGES_TO_DELETE:
            for message_id in MESSAGES_TO_DELETE[chat_id]:
                try:
                    await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
                except Exception as e:
                    logger.debug(f"Не удалось удалить сообщение {message_id}: {e}")
            
            MESSAGES_TO_DELETE[chat_id] = []
            
        logger.info(f"История чата очищена для пользователя {update.effective_user.id}")
    except Exception as e:
        logger.error(f"Ошибка при очистке истории чата: {e}")

async def send_message_with_cleanup(update: Update, context: ContextTypes.DEFAULT_TYPE, text, **kwargs):
    """Отправка сообщения с автоматическим добавлением в список для удаления"""
    message = await update.message.reply_text(text, **kwargs)
    await add_message_to_delete(update.effective_chat.id, message.message_id)
    return message

def optimize_stringers(stringer_length):
    """Оптимизация раскроя тетивы для минимизации отходов"""
    total_stringer_qty = 2
    
    if stringer_length <= 3000:
        return [{'length': 3000, 'qty': total_stringer_qty}], total_stringer_qty
    
    elif stringer_length <= 4000:
        return [{'length': 4000, 'qty': total_stringer_qty}], total_stringer_qty
    
    else:
        qty_4000 = math.ceil(stringer_length / 4000) * total_stringer_qty
        waste_4000 = (qty_4000 * 4000) - (stringer_length * total_stringer_qty)
        
        qty_4000_combo = math.floor(stringer_length / 4000) * total_stringer_qty
        remaining_length = (stringer_length * total_stringer_qty) - (qty_4000_combo * 4000)
        qty_3000_combo = math.ceil(remaining_length / 3000) if remaining_length > 0 else 0
        waste_combo = (qty_4000_combo * 4000 + qty_3000_combo * 3000) - (stringer_length * total_stringer_qty)
        
        if waste_4000 <= waste_combo:
            return [{'length': 4000, 'qty': qty_4000}], qty_4000
        else:
            result = []
            if qty_4000_combo > 0:
                result.append({'length': 4000, 'qty': qty_4000_combo})
            if qty_3000_combo > 0:
                result.append({'length': 3000, 'qty': qty_3000_combo})
            return result, qty_4000_combo + qty_3000_combo

def calculate_wood_stairs(height, steps_count, config, material_type, actual_step_height, step_width):
    """Расчет деревянной лестницы с фиксированной высотой ступени 225 мм"""
    materials = []
    total_cost = 0
    
    steps_count = math.ceil(height / FIXED_STEP_HEIGHT)
    actual_step_height = height / steps_count
    
    platforms_count = 0
    if config == 'l_shape':
        platforms_count = 1
        steps_count = max(1, steps_count - 1)
    elif config == 'u_shape':
        platforms_count = 2
        steps_count = max(1, steps_count - 2)
    
    step_depth = 300
    
    if config == 'straight':
        stair_length = (steps_count - 1) * step_depth
        stringer_length = math.sqrt(height**2 + stair_length**2)
        total_stringer_length = stringer_length * 2
        
    elif config == 'l_shape':
        first_flight_steps = math.ceil(steps_count / 2)
        second_flight_steps = steps_count - first_flight_steps
        
        first_flight_height = first_flight_steps * actual_step_height
        second_flight_height = second_flight_steps * actual_step_height
        
        first_flight_length = (first_flight_steps - 1) * step_depth
        second_flight_length = (second_flight_steps - 1) * step_depth
        
        first_stringer_length = math.sqrt(first_flight_height**2 + first_flight_length**2)
        second_stringer_length = math.sqrt(second_flight_height**2 + second_flight_length**2)
        
        total_stringer_length = (first_stringer_length + second_stringer_length) * 2
        
    else:
        flights_steps = math.ceil(steps_count / 3)
        remaining_steps = steps_count - flights_steps * 2
        if remaining_steps < 0:
            flights_steps = math.ceil(steps_count / 2)
            remaining_steps = steps_count - flights_steps
        
        flight_height = flights_steps * actual_step_height
        flight_length = (flights_steps - 1) * step_depth
        
        flight_stringer_length = math.sqrt(flight_height**2 + flight_length**2)
        total_stringer_length = flight_stringer_length * 4
    
    stringers_optimized, total_stringer_qty = optimize_stringers(total_stringer_length / 2)
    
    for stringer in stringers_optimized:
        stringer_price = get_material_price(material_type, f'Тетива {stringer["length"]}', 10215 if stringer["length"] == 4000 else 9518)
        stringer_cost = stringer_price * stringer["qty"]
        
        materials.append({
            'name': f'Тетива {stringer["length"]}мм',
            'qty': stringer["qty"],
            'unit': 'шт.',
            'price': stringer_price,
            'total': stringer_cost
        })
        total_cost += stringer_cost
    
    step_price = get_material_price(material_type, f'СТУПЕНЬ ПРЯМАЯ {step_width}', 1500)
    step_cost = steps_count * step_price
    
    materials.append({
        'name': f'Ступень {step_width}×300мм',
        'qty': steps_count,
        'unit': 'шт.',
        'price': step_price,
        'total': step_cost
    })
    total_cost += step_cost
    
    riser_price = get_material_price(material_type, f'Подступенок {step_width}', 600)
    riser_cost = steps_count * riser_price
    
    materials.append({
        'name': f'Подступенок {step_width}×200мм',
        'qty': steps_count,
        'unit': 'шт.',
        'price': riser_price,
        'total': riser_cost
    })
    total_cost += riser_cost
    
    if platforms_count > 0:
        platform_size = 1000 if step_width in ["900", "1000"] else 1200
        platform_price = get_material_price(material_type, f'Площадка {platform_size}', 8000 if platform_size == 1000 else 9500)
        platform_cost = platforms_count * platform_price
        
        materials.append({
            'name': f'Площадка {platform_size}×{platform_size}мм',
            'qty': platforms_count,
            'unit': 'шт.',
            'price': platform_price,
            'total': platform_cost
        })
        total_cost += platform_cost
    
    post_price = get_material_price(material_type, 'Столб', 1931)
    if config == 'straight':
        posts_qty = 2
    elif config == 'l_shape':
        posts_qty = 3
    else:
        posts_qty = 4
    
    posts_cost = posts_qty * post_price
    
    materials.append({
        'name': 'Столб опорный',
        'qty': posts_qty,
        'unit': 'шт.',
        'price': post_price,
        'total': posts_cost
    })
    total_cost += posts_cost
    
    baluster_price = get_material_price(material_type, 'Балясина', 400)
    balusters_qty = steps_count + platforms_count
    balusters_cost = balusters_qty * baluster_price
    
    materials.append({
        'name': 'Балясина',
        'qty': balusters_qty,
        'unit': 'шт.',
        'price': baluster_price,
        'total': balusters_cost
    })
    total_cost += balusters_cost
    
    handrail_length = total_stringer_length / 2
    handrail_qty = math.ceil(handrail_length / 3000)
    handrail_price = get_material_price(material_type, 'ПОРУЧЕНЬ', 2108)
    handrail_cost = handrail_qty * handrail_price
    
    materials.append({
        'name': 'Поручень 3000мм',
        'qty': handrail_qty,
        'unit': 'шт.',
        'price': handrail_price,
        'total': handrail_cost
    })
    total_cost += handrail_cost
    
    return {
        'type': 'wood',
        'config': config,
        'height': height,
        'step_width': step_width,
        'steps_count': steps_count,
        'platforms_count': platforms_count,
        'step_height': actual_step_height,
        'stringer_length': total_stringer_length / 2,
        'stringer_qty': total_stringer_qty,
        'stringers_detail': stringers_optimized,
        'posts_count': posts_qty,
        'materials': materials,
        'total_cost': total_cost
    }

def calculate_modular_stairs(height, steps_count, config, material_type, actual_step_height, step_width):
    """Расчет модульной лестницы с фиксированной высотой ступени 225 мм"""
    materials = []
    total_cost = 0
    
    steps_count = math.ceil(height / FIXED_STEP_HEIGHT)
    actual_step_height = height / steps_count
    
    platforms_count = 0
    if config == 'l_shape':
        platforms_count = 1
        steps_count = max(1, steps_count - 1)
    elif config == 'u_shape':
        platforms_count = 2
        steps_count = max(1, steps_count - 2)
    
    support_1000 = get_material_by_article('15762374')
    support_2000 = get_material_by_article('15762382')
    
    if support_1000:
        materials.append({
            'name': support_1000['name'],
            'qty': 1,
            'unit': 'шт.',
            'price': support_1000['price'],
            'total': support_1000['price']
        })
        total_cost += support_1000['price']
    
    if support_2000:
        materials.append({
            'name': support_2000['name'],
            'qty': 1,
            'unit': 'шт.',
            'price': support_2000['price'],
            'total': support_2000['price']
        })
        total_cost += support_2000['price']
    
    module_price = get_material_price(material_type, 'Промежуточный элемент', 4076)
    modules_qty = steps_count - 1
    modules_cost = modules_qty * module_price
    
    materials.append({
        'name': 'Промежуточный элемент',
        'qty': modules_qty,
        'unit': 'шт.',
        'price': module_price,
        'total': modules_cost
    })
    total_cost += modules_cost
    
    end_module_price = get_material_price(material_type, 'Верхний и нижний элемент', 7590)
    materials.append({
        'name': 'Верхний и нижний элемент',
        'qty': 1,
        'unit': 'шт.',
        'price': end_module_price,
        'total': end_module_price
    })
    total_cost += end_module_price
    
    corner_element = get_material_by_article('15762391')
    if corner_element:
        if config == 'l_shape':
            materials.append({
                'name': corner_element['name'],
                'qty': 1,
                'unit': 'шт.',
                'price': corner_element['price'],
                'total': corner_element['price']
            })
            total_cost += corner_element['price']
        elif config == 'u_shape':
            materials.append({
                'name': corner_element['name'],
                'qty': 2,
                'unit': 'шт.',
                'price': corner_element['price'],
                'total': corner_element['price'] * 2
            })
            total_cost += corner_element['price'] * 2
    
    if platforms_count > 0:
        platform_price = get_material_price(material_type, 'Площадка', 8000)
        materials.append({
            'name': f'Площадка {step_width}x{step_width}',
            'qty': platforms_count,
            'unit': 'шт.',
            'price': platform_price,
            'total': platform_price * platforms_count
        })
        total_cost += platform_price * platforms_count
    
    step_price = get_material_price(material_type, f'СТУПЕНЬ ПРЯМАЯ {step_width}', 1500)
    step_cost = steps_count * step_price
    
    materials.append({
        'name': f'Ступень {step_width}×300мм',
        'qty': steps_count,
        'unit': 'шт.',
        'price': step_price,
        'total': step_cost
    })
    total_cost += step_cost
    
    railing_price = get_material_price(material_type, 'Опора под поручень', 900)
    railing_qty = steps_count + platforms_count
    railing_cost = railing_qty * railing_price
    
    materials.append({
        'name': 'Опора под поручень',
        'qty': railing_qty,
        'unit': 'шт.',
        'price': railing_price,
        'total': railing_cost
    })
    total_cost += railing_cost
    
    handrail_length = math.sqrt(height**2 + (steps_count * 300)**2) / 1000
    handrail_qty = math.ceil(handrail_length / 3)
    handrail_price = get_material_price(material_type, 'ПОРУЧЕНЬ', 2108)
    handrail_cost = handrail_qty * handrail_price
    
    materials.append({
        'name': 'Поручень 3000мм',
        'qty': handrail_qty,
        'unit': 'шт.',
        'price': handrail_price,
        'total': handrail_cost
    })
    total_cost += handrail_cost
    
    return {
        'type': 'modular',
        'config': config,
        'height': height,
        'step_width': step_width,
        'steps_count': steps_count,
        'platforms_count': platforms_count,
        'step_height': actual_step_height,
        'stringer_length': handrail_length,
        'materials': materials,
        'total_cost': total_cost
    }

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    await cleanup_chat_history(update, context)
    
    if prices_data is None:
        load_prices()
    
    user = update.effective_user
    welcome_text = (
        f"👋 Добро пожаловать, {user.first_name}!\n"
        "Я твой помощник в расчете лестниц.\n\n"
        "📋 *Доступные функции:*\n"
        "• 🏠 *Расчет лестницы* - полный расчет стоимости\n"
        "• 🔍 *Поиск материала* - найти материал по артикулу или названию\n\n"
        "Выберите действие:"
    )
    
    keyboard = [
        [InlineKeyboardButton("🔄 Рассчитать лестницу", callback_data="calculate_stairs")],
        [InlineKeyboardButton("🔍 Найти материал", callback_data="search_material")],
        [InlineKeyboardButton("🔄 Перезапустить", callback_data="restart")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    message = await update.message.reply_text(welcome_text, reply_markup=reply_markup, parse_mode='Markdown')
    await add_message_to_delete(update.effective_chat.id, message.message_id)

async def restart_bot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Перезапуск бота"""
    query = update.callback_query
    await query.answer()
    
    chat_id = query.message.chat_id
    if chat_id in MESSAGES_TO_DELETE:
        for message_id in MESSAGES_TO_DELETE[chat_id]:
            try:
                await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
            except Exception as e:
                logger.debug(f"Не удалось удалить сообщение {message_id}: {e}")
        MESSAGES_TO_DELETE[chat_id] = []
    
    user = query.from_user
    user_id = user.id
    if user_id in user_data:
        del user_data[user_id]
    
    welcome_text = (
        f"👋 Добро пожаловать, {user.first_name}!\n"
        "Я твой помощник в расчете лестниц.\n\n"
        "📋 *Доступные функции:*\n"
        "• 🏠 *Расчет лестницы* - полный расчет стоимости\n"
        "• 🔍 *Поиск материала* - найти материал по артикулу или названию\n\n"
        "Выберите действие:"
    )
    
    keyboard = [
        [InlineKeyboardButton("🔄 Рассчитать лестницу", callback_data="calculate_stairs")],
        [InlineKeyboardButton("🔍 Найти материал", callback_data="search_material")],
        [InlineKeyboardButton("🔄 Перезапустить", callback_data="restart")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(welcome_text, reply_markup=reply_markup, parse_mode='Markdown')

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик нажатий на кнопки"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "calculate_stairs":
        await cleanup_chat_history(update, context)
        
        user_id = query.from_user.id
        if user_id not in user_data:
            user_data[user_id] = {}
        
        reply_keyboard = [
            ["🏠 Деревянная", "⚡ Модульная"],
            ["🔍 Найти материал", "🔄 Перезапустить"]
        ]
        
        message = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="👋 Добро пожаловать!\n\n"
                 "📋 *Выберите тип лестницы:*\n"
                 "• 🏠 *Деревянная* - из отдельных элементов\n"
                 "• ⚡ *Модульная* - металлическая система",
            reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True),
            parse_mode='Markdown'
        )
        await add_message_to_delete(query.message.chat_id, message.message_id)
        return SELECTING_TYPE
    
    elif query.data == "search_material":
        await cleanup_chat_history(update, context)
        
        reply_keyboard = [["🔄 Перезапустить"]]
        
        message = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="🔍 *ПОИСК МАТЕРИАЛА*\n\n"
                 "Введите артикул или название материала для поиска:\n\n"
                 "Примеры:\n"
                 "• `15762294` - поиск по артикулу\n"
                 "• `Ступень` - поиск по названию\n"
                 "• `Тетива` - поиск по названию",
            reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True),
            parse_mode='Markdown'
        )
        await add_message_to_delete(query.message.chat_id, message.message_id)
        return SEARCH_MATERIAL
    
    elif query.data == "restart":
        await restart_bot(update, context)

async def search_material(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Поиск материала по артикулу или названию"""
    search_term = update.message.text
    user_id = update.effective_user.id
    
    await add_message_to_delete(update.effective_chat.id, update.message.message_id)
    
    if search_term == "🔄 Перезапустить":
        await restart_from_message(update, context)
        return ConversationHandler.END
    
    if not search_term.strip():
        await send_message_with_cleanup(update, context, "❌ Пожалуйста, введите артикул или название для поиска")
        return SEARCH_MATERIAL
    
    search_msg = await send_message_with_cleanup(update, context, "🔍 Ищу материалы...")
    
    results = search_materials_by_article_or_name(search_term)
    
    try:
        await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=search_msg.message_id)
    except:
        pass
    
    if not results:
        await send_message_with_cleanup(
            update, context,
            f"❌ Материалы по запросу '{search_term}' не найдены.\n\n"
            "Попробуйте:\n"
            "• Проверить правильность артикула\n"
            "• Использовать другое название\n"
            "• Упростить запрос"
        )
        return SEARCH_MATERIAL
    
    message_text = f"🔍 *РЕЗУЛЬТАТЫ ПОИСКА* ('{search_term}')\n\n"
    
    for i, item in enumerate(results[:10], 1):  # Ограничиваем вывод 10 результатами
        message_text += (
            f"*{i}. {item['name']}*\n"
            f"📋 Артикул: `{item['article']}`\n"
            f"🏷 Тип: {item['stair_type']}\n"
            f"📏 Размеры: {item.get('sizes', 'не указаны')}\n"
            f"💰 Цена: {item['price']:,.0f} ₽\n"
            f"📦 Ед. изм.: {item['unit']}\n\n"
        )
    
    if len(results) > 10:
        message_text += f"*... и еще {len(results) - 10} материалов*"
    
    message_text += "\n_Для нового поиска введите артикул или название_"
    
    keyboard = [
        [InlineKeyboardButton("🔄 Новый поиск", callback_data="search_material")],
        [InlineKeyboardButton("🏠 Расчет лестницы", callback_data="calculate_stairs")],
        [InlineKeyboardButton("🔄 Перезапустить", callback_data="restart")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    message = await update.message.reply_text(message_text, reply_markup=reply_markup, parse_mode='Markdown')
    await add_message_to_delete(update.effective_chat.id, message.message_id)
    
    return ConversationHandler.END

async def select_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Выбор типа лестницы"""
    user_choice = update.message.text
    user_id = update.effective_user.id
    
    await add_message_to_delete(update.effective_chat.id, update.message.message_id)
    
    if user_choice == "🔄 Перезапустить":
        await restart_from_message(update, context)
        return ConversationHandler.END
    
    if user_choice == "🔍 Найти материал":
        await cleanup_chat_history(update, context)
        
        reply_keyboard = [["🔄 Перезапустить"]]
        
        await send_message_with_cleanup(
            update, context,
            "🔍 *ПОИСК МАТЕРИАЛА*\n\n"
            "Введите артикул или название материала для поиска:\n\n"
            "Примеры:\n"
            "• `15762294` - поиск по артикулу\n"
            "• `Ступень` - поиск по названию\n"
            "• `Тетива` - поиск по названию",
            reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True),
            parse_mode='Markdown'
        )
        return SEARCH_MATERIAL
    
    user_data[user_id] = {
        'type': 'wood' if 'Деревянная' in user_choice else 'modular',
        'material_type': 'деревянная' if 'Деревянная' in user_choice else 'металлическая'
    }
    
    reply_keyboard = [
        ["📏 Прямая", "📐 Г-образная", "🔄 П-образная"],
        ["🔍 Найти материал", "🔄 Перезапустить"]
    ]
    
    await send_message_with_cleanup(
        update, context,
        "📐 *Выберите конфигурацию лестницы:*\n\n"
        "• 📏 *Прямая* - одномаршевая лестница\n"
        "• 📐 *Г-образная* - с поворотом на 90°\n" 
        "• 🔄 *П-образная* - с поворотом на 180°",
        reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True),
        parse_mode='Markdown'
    )
    return SELECTING_CONFIG

async def select_config(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Выбор конфигурации лестницы"""
    user_choice = update.message.text
    user_id = update.effective_user.id
    
    await add_message_to_delete(update.effective_chat.id, update.message.message_id)
    
    if user_choice == "🔄 Перезапустить":
        await restart_from_message(update, context)
        return ConversationHandler.END
    
    if user_choice == "🔍 Найти материал":
        await cleanup_chat_history(update, context)
        
        reply_keyboard = [["🔄 Перезапустить"]]
        
        await send_message_with_cleanup(
            update, context,
            "🔍 *ПОИСК МАТЕРИАЛА*\n\n"
            "Введите артикул или название материала для поиска:\n\n"
            "Примеры:\n"
            "• `15762294` - поиск по артикулу\n"
            "• `Ступень` - поиск по названию\n"
            "• `Тетива` - поиск по названию",
            reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True),
            parse_mode='Markdown'
        )
        return SEARCH_MATERIAL
    
    config_map = {
        '📏 Прямая': 'straight',
        '📐 Г-образная': 'l_shape', 
        '🔄 П-образная': 'u_shape'
    }
    
    user_data[user_id]['config'] = config_map[user_choice]
    
    reply_keyboard = [["🔄 Перезапустить"]]
    
    await send_message_with_cleanup(
        update, context,
        "📏 *Введите высоту лестницы (мм):*\n\n"
        "Примеры:\n"
        "• 2700 - для высоты 2.7 метра\n" 
        "• 3000 - для высоты 3 метра\n"
        "• 3500 - для высоты 3.5 метра\n\n"
        "📝 *Рекомендация:* Высота измеряется от чистого пола нижнего этажа до чистого пола верхнего этажа",
        reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True),
        parse_mode='Markdown'
    )
    return INPUT_HEIGHT

async def input_height(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Ввод высоты лестницы"""
    height_input = update.message.text
    user_id = update.effective_user.id
    
    await add_message_to_delete(update.effective_chat.id, update.message.message_id)
    
    if height_input == "🔄 Перезапустить":
        await restart_from_message(update, context)
        return INPUT_HEIGHT
    
    is_valid, result = validate_input(height_input, 1000, 5000, "Высота лестницы")
    
    if not is_valid:
        await send_message_with_cleanup(update, context, result)
        return INPUT_HEIGHT
    
    user_data[user_id]['height'] = result
    
    reply_keyboard = [
        ["900", "1000", "1200"],
        ["🔍 Найти материал", "🔄 Перезапустить"]
    ]
    
    await send_message_with_cleanup(
        update, context,
        "📐 *Выберите ширину ступени:*\n\n"
        "• 900 мм - компактный вариант\n"
        "• 1000 мм - стандартная ширина\n"
        "• 1200 мм - просторная лестница",
        reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True),
        parse_mode='Markdown'
    )
    return SELECTING_STEP_SIZE

async def select_step_size(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Выбор ширины ступени"""
    step_width = update.message.text
    user_id = update.effective_user.id
    
    await add_message_to_delete(update.effective_chat.id, update.message.message_id)
    
    if step_width == "🔄 Перезапустить":
        await restart_from_message(update, context)
        return ConversationHandler.END
    
    if step_width == "🔍 Найти материал":
        await cleanup_chat_history(update, context)
        
        reply_keyboard = [["🔄 Перезапустить"]]
        
        await send_message_with_cleanup(
            update, context,
            "🔍 *ПОИСК МАТЕРИАЛА*\n\n"
            "Введите артикул или название материала для поиска:\n\n"
            "Примеры:\n"
            "• `15762294` - поиск по артикулу\n"
            "• `Ступень` - поиск по названию\n"
            "• `Тетива` - поиск по названию",
            reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True),
            parse_mode='Markdown'
        )
        return SEARCH_MATERIAL
    
    if step_width not in ["900", "1000", "1200"]:
        await send_message_with_cleanup(update, context, "❌ Пожалуйста, выберите ширину ступени из предложенных вариантов")
        return SELECTING_STEP_SIZE
    
    user_data[user_id]['step_width'] = step_width
    
    calculation_msg = await send_message_with_cleanup(update, context, "🧮 *Выполняю расчет...*", parse_mode='Markdown')
    
    user_input = user_data[user_id]
    
    try:
        if user_input['type'] == 'wood':
            result = calculate_wood_stairs(
                height=user_input['height'],
                steps_count=0,
                config=user_input['config'],
                material_type=user_input['material_type'],
                actual_step_height=FIXED_STEP_HEIGHT,
                step_width=user_input['step_width']
            )
        else:
            result = calculate_modular_stairs(
                height=user_input['height'],
                steps_count=0,
                config=user_input['config'],
                material_type=user_input['material_type'],
                actual_step_height=FIXED_STEP_HEIGHT,
                step_width=user_input['step_width']
            )
        
        await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=calculation_msg.message_id)
        
        config_names = {
            'straight': 'Прямая',
            'l_shape': 'Г-образная', 
            'u_shape': 'П-образная'
        }
        
        type_names = {
            'wood': 'Деревянная',
            'modular': 'Модульная'
        }
        
        result_text = (
            f"📊 *РЕЗУЛЬТАТ РАСЧЕТА*\n\n"
            f"🏷 *Тип:* {type_names[result['type']]}\n"
            f"📐 *Конфигурация:* {config_names[result['config']]}\n"
            f"📏 *Высота:* {result['height']:,} мм\n"
            f"📐 *Ширина ступени:* {result['step_width']} мм\n"
            f"🪜 *Количество ступеней:* {result['steps_count']}\n"
            f"📏 *Высота ступени:* {result['step_height']:.1f} мм\n"
        )
        
        if result['type'] == 'wood':
            result_text += f"📏 *Длина тетивы:* {result['stringer_length']:.0f} мм\n"
            result_text += f"🔢 *Количество тетив:* {result['stringer_qty']} шт\n"
        
        if result['platforms_count'] > 0:
            result_text += f"🟦 *Количество площадок:* {result['platforms_count']}\n"
        
        result_text += f"\n📦 *МАТЕРИАЛЫ:*\n"
        
        for material in result['materials']:
            result_text += f"• {material['name']}: {material['qty']} {material['unit']} × {material['price']:,.0f} ₽ = {material['total']:,.0f} ₽\n"
        
        result_text += f"\n💰 *ОБЩАЯ СТОИМОСТЬ:* {result['total_cost']:,.0f} ₽\n\n"
        result_text += "_*Примечание:* Стоимость указана без учета доставки и монтажа_\n"
        
        keyboard = [
            [InlineKeyboardButton("🔄 Новый расчет", callback_data="calculate_stairs")],
            [InlineKeyboardButton("🔍 Поиск материала", callback_data="search_material")],
            [InlineKeyboardButton("🔄 Перезапустить", callback_data="restart")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        message = await update.message.reply_text(result_text, reply_markup=reply_markup, parse_mode='Markdown')
        await add_message_to_delete(update.effective_chat.id, message.message_id)
        
        return ConversationHandler.END
        
    except Exception as e:
        logger.error(f"Ошибка расчета: {e}")
        await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=calculation_msg.message_id)
        await send_message_with_cleanup(update, context, f"❌ Произошла ошибка при расчете: {str(e)}")
        return ConversationHandler.END

async def restart_from_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Перезапуск из состояния диалога"""
    await cleanup_chat_history(update, context)
    
    user_id = update.effective_user.id
    if user_id in user_data:
        del user_data[user_id]
    
    user = update.effective_user
    welcome_text = (
        f"👋 Добро пожаловать, {user.first_name}!\n"
        "Я твой помощник в расчете лестниц.\n\n"
        "📋 *Доступные функции:*\n"
        "• 🏠 *Расчет лестницы* - полный расчет стоимости\n"
        "• 🔍 *Поиск материала* - найти материал по артикулу или названию\n\n"
        "Выберите действие:"
    )
    
    keyboard = [
        [InlineKeyboardButton("🔄 Рассчитать лестницу", callback_data="calculate_stairs")],
        [InlineKeyboardButton("🔍 Найти материал", callback_data="search_material")],
        [InlineKeyboardButton("🔄 Перезапустить", callback_data="restart")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    message = await update.message.reply_text(welcome_text, reply_markup=reply_markup, parse_mode='Markdown')
    await add_message_to_delete(update.effective_chat.id, message.message_id)

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Отмена диалога"""
    await cleanup_chat_history(update, context)
    
    user_id = update.effective_user.id
    if user_id in user_data:
        del user_data[user_id]
    
    await send_message_with_cleanup(update, context, "Диалог отменен. Используйте /start для начала нового расчета.")
    return ConversationHandler.END

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ошибок"""
    logger.error(f"Ошибка: {context.error}", exc_info=context.error)
    
    try:
        await send_message_with_cleanup(update, context, "❌ Произошла ошибка. Пожалуйста, попробуйте еще раз или используйте /start для перезапуска.")
    except:
        pass

def main():
    """Основная функция запуска бота"""
    # Запускаем keep-alive сервер для Replit
    keep_alive()
    logger.info("🔄 Keep-alive server started on port 8080")
    
    # Опционально: запускаем само-пинг (раскомментируйте если нужно)
    # start_ping_loop()
    # logger.info("🔁 Self-ping service started")
    
    # Загружаем цены при старте
    load_prices()
    
    # Получаем токен
    token = os.getenv('TELEGRAM_BOT_TOKEN')
    if not token:
        logger.error("❌ TELEGRAM_BOT_TOKEN не найден в Secrets")
        logger.info("📝 Добавьте TELEGRAM_BOT_TOKEN в раздел Secrets (Tools → Secrets)")
        return
    
    # Создаем приложение
    application = Application.builder().token(token).build()
    
    # Обработчик диалога
    conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler('start', start),
            CallbackQueryHandler(button_handler, pattern='^(calculate_stairs|search_material)$')
        ],
        states={
            SELECTING_TYPE: [MessageHandler(filters.TEXT & ~filters.COMMAND, select_type)],
            SELECTING_CONFIG: [MessageHandler(filters.TEXT & ~filters.COMMAND, select_config)],
            INPUT_HEIGHT: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_height)],
            SELECTING_STEP_SIZE: [MessageHandler(filters.TEXT & ~filters.COMMAND, select_step_size)],
            SEARCH_MATERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, search_material)],
        },
        fallbacks=[
            CommandHandler('cancel', cancel),
            CommandHandler('start', start),
            CallbackQueryHandler(restart_bot, pattern='^restart$'),
            CallbackQueryHandler(button_handler, pattern='^(calculate_stairs|search_material)$')
        ],
        allow_reentry=True
    )
    
    application.add_handler(conv_handler)
    application.add_handler(CallbackQueryHandler(restart_bot, pattern='^restart$'))
    application.add_error_handler(error_handler)
    
    logger.info("🚀 Бот запущен и готов к работе!")
    logger.info("📡 Keep-alive сервер работает на порту 8080")
    logger.info("🔗 URL для мониторинга: https://your-repl-name.your-username.repl.co")
    
    application.run_polling(drop_pending_updates=True)

if __name__ == '__main__':
    main()
